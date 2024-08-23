import os
from openpyxl import Workbook, load_workbook, utils

def update_spreadsheet(data, path):
    if os.path.exists(path):
        # Spreadsheet exists then access it to update data
        wb = load_workbook(path)

        ws_name = "2024"
        if ws_name not in wb.sheetnames:
            raise Exception("Worksheet does not exist")
        ws = wb[ws_name]
    else:
        raise Exception("Spreadsheet does not exist")

    # Find first empty row in worksheet
    end_column = utils.column_index_from_string('K')

    for row in ws.iter_rows(min_row=1, max_col=end_column, max_row=ws.max_row):
        if row[0].value is None:
            row_index = row[0].row
            ws[f'A{row_index}'] = data["First Name"]
            ws[f'B{row_index}'] = data["Last Name"]
            ws[f'C{row_index}'] = data["ID"]
            ws[f'D{row_index}'] = data["UPI"]
            ws[f'E{row_index}'] = data["Card Number"]
            ws[f'F{row_index}'] = data["Programme"]
            ws[f'G{row_index}'] = data["Level of Appointment"]
            ws[f'H{row_index}'] = data["Access Group"]
            # ws[f'I{row_index}'] = data["Request Sent"]
            # ws[f'J{row_index}'] = data["Contract Expiry"]
            break

    # Save the spreadsheet
    wb.save(path)
    return f"Data added to row {row_index}"